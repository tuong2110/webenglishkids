#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script convert file Excel ETS sang CSV
Format CSV: Test,Part,NoiDungTuVung,DichNghia,AmThanh,CauViDu
"""

import sys
import os
import csv

try:
    import openpyxl
except ImportError:
    print("❌ Chưa cài đặt thư viện openpyxl!")
    print("📦 Cài đặt bằng lệnh: pip install openpyxl")
    sys.exit(1)

input_file = "ETS 2024 - LISTENING.xlsx"
output_file = "ETS_2024_LISTENING.csv"

if not os.path.exists(input_file):
    print(f"❌ Không tìm thấy file: {input_file}")
    print("💡 Đảm bảo file nằm trong thư mục gốc của project")
    sys.exit(1)

try:
    print(f"📖 Đang đọc file: {input_file}")
    print(f"📝 Sẽ xuất ra: {output_file}\n")
    
    # Mở file Excel
    wb = openpyxl.load_workbook(input_file, data_only=True)
    
    # Tạo file CSV
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        
        # Ghi header
        writer.writerow(['Test', 'Part', 'NoiDungTuVung', 'DichNghia', 'AmThanh', 'CauViDu'])
        
        total_rows = 0
        
        # Đọc từng sheet (mỗi sheet = 1 test)
        for sheet_index, sheet_name in enumerate(wb.sheetnames, 1):
            sheet = wb[sheet_name]
            print(f"📄 Đang xử lý sheet: {sheet_name} (Test {sheet_index})")
            
            # Đọc header (dòng 1)
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Column{col}")
            
            # Tìm index các cột
            idx_phan_loai = None
            idx_tu_tieng_anh = None
            idx_nghia_tieng_viet = None
            idx_anh_my = None
            idx_cau_vi_du = None
            
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if 'phân loại' in header_lower or 'phan loai' in header_lower:
                    idx_phan_loai = idx + 1
                if 'từ tiếng anh' in header_lower or 'tu tieng anh' in header_lower:
                    idx_tu_tieng_anh = idx + 1
                if 'nghĩa tiếng việt' in header_lower or 'nghia tieng viet' in header_lower:
                    idx_nghia_tieng_viet = idx + 1
                if 'anh - mỹ' in header_lower or 'anh - my' in header_lower or 'anh mỹ' in header_lower:
                    idx_anh_my = idx + 1
                if 'câu ví dụ' in header_lower or 'cau vi du' in header_lower:
                    idx_cau_vi_du = idx + 1
            
            if not idx_phan_loai or not idx_tu_tieng_anh or not idx_nghia_tieng_viet:
                print(f"   ⚠️ Sheet '{sheet_name}': Thiếu cột bắt buộc, bỏ qua")
                continue
            
            # Đọc từng dòng (bắt đầu từ dòng 2)
            sheet_rows = 0
            for row_idx in range(2, sheet.max_row + 1):
                # Lấy giá trị từ các cột
                phan_loai = sheet.cell(row=row_idx, column=idx_phan_loai).value
                tu_tieng_anh = sheet.cell(row=row_idx, column=idx_tu_tieng_anh).value
                nghia_tieng_viet = sheet.cell(row=row_idx, column=idx_nghia_tieng_viet).value
                anh_my = sheet.cell(row=row_idx, column=idx_anh_my).value if idx_anh_my else None
                cau_vi_du = sheet.cell(row=row_idx, column=idx_cau_vi_du).value if idx_cau_vi_du else None
                
                # Bỏ qua dòng trống
                if not tu_tieng_anh or not nghia_tieng_viet:
                    continue
                
                # Xử lý Part từ "Phân loại" (có thể là "Parrt 1", "Part 1", "Part 2", etc.)
                part_number = 1  # Mặc định
                if phan_loai:
                    phan_loai_str = str(phan_loai).strip()
                    # Tìm số trong "Parrt 1", "Part 1", "Part 2", etc.
                    import re
                    match = re.search(r'part\s*(\d+)', phan_loai_str, re.IGNORECASE)
                    if match:
                        part_number = int(match.group(1))
                
                # Chuẩn hóa dữ liệu
                test_number = sheet_index
                tu_tieng_anh_str = str(tu_tieng_anh).strip() if tu_tieng_anh else ''
                nghia_tieng_viet_str = str(nghia_tieng_viet).strip() if nghia_tieng_viet else ''
                anh_my_str = str(anh_my).strip() if anh_my else ''
                cau_vi_du_str = str(cau_vi_du).strip() if cau_vi_du else ''
                
                # Ghi vào CSV
                writer.writerow([
                    test_number,
                    part_number,
                    tu_tieng_anh_str,
                    nghia_tieng_viet_str,
                    anh_my_str,
                    cau_vi_du_str
                ])
                
                sheet_rows += 1
                total_rows += 1
            
            print(f"   ✅ Đã xử lý {sheet_rows} từ vựng từ {sheet_name}")
        
        print(f"\n✅ Hoàn tất!")
        print(f"📊 Tổng cộng: {total_rows} từ vựng")
        print(f"📁 File CSV: {output_file}")
        print(f"\n💡 Bây giờ bạn có thể:")
        print(f"   1. Sử dụng script import CSV: http://localhost:8000/import_tuvung_csv.php?confirm=yes")
        print(f"   2. Bật checkbox 'Tự động tạo bài học từ Part'")
        print(f"   3. Chọn khóa học và upload file {output_file}")
        
except Exception as e:
    print(f"❌ Lỗi: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)


