#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script đọc file Excel ETS và hiển thị cấu trúc dữ liệu
"""

import sys
import os

try:
    import openpyxl
except ImportError:
    print("❌ Chưa cài đặt thư viện openpyxl!")
    print("📦 Cài đặt bằng lệnh: pip install openpyxl")
    sys.exit(1)

file_path = "ETS 2024 - LISTENING.xlsx"

if not os.path.exists(file_path):
    print(f"❌ Không tìm thấy file: {file_path}")
    print("💡 Đảm bảo file nằm trong thư mục gốc của project")
    sys.exit(1)

try:
    print(f"📖 Đang đọc file: {file_path}\n")
    print("=" * 80)
    
    # Mở file Excel
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Liệt kê các sheet
    print(f"\n📋 Có {len(wb.sheetnames)} sheet(s):")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"   {idx}. {sheet_name}")
    
    # Đọc sheet đầu tiên
    sheet = wb.active
    print(f"\n📄 Đang đọc sheet: '{sheet.title}'")
    print(f"   - Số dòng: {sheet.max_row}")
    print(f"   - Số cột: {sheet.max_column}")
    
    # Đọc header (dòng đầu tiên)
    print("\n" + "=" * 80)
    print("📊 HEADER (Dòng đầu tiên):")
    print("=" * 80)
    
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value:
            headers.append(str(cell_value).strip())
        else:
            headers.append(f"Column{col}")
    
    for idx, header in enumerate(headers, 1):
        print(f"   {idx}. {header}")
    
    # Đọc 10 dòng đầu tiên (sau header)
    print("\n" + "=" * 80)
    print("📝 10 DÒNG ĐẦU TIÊN (sau header):")
    print("=" * 80)
    
    max_rows_to_show = min(10, sheet.max_row - 1)
    
    for row_idx in range(2, 2 + max_rows_to_show):
        row_data = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row_idx, column=col).value
            if cell_value is None:
                row_data.append("")
            else:
                row_data.append(str(cell_value).strip())
        
        # Chỉ hiển thị nếu dòng không rỗng
        if any(row_data):
            print(f"\nDòng {row_idx}:")
            for col_idx, value in enumerate(row_data, 1):
                if value:  # Chỉ hiển thị cột có giá trị
                    print(f"   {headers[col_idx-1]}: {value}")
    
    # Phân tích cấu trúc
    print("\n" + "=" * 80)
    print("🔍 PHÂN TÍCH CẤU TRÚC:")
    print("=" * 80)
    
    # Tìm cột Test và Part
    test_col = None
    part_col = None
    word_col = None
    meaning_col = None
    
    for idx, header in enumerate(headers):
        header_lower = header.lower()
        if 'test' in header_lower:
            test_col = idx + 1
        if 'part' in header_lower:
            part_col = idx + 1
        if any(x in header_lower for x in ['word', 'tuvung', 'noidung']):
            word_col = idx + 1
        if any(x in header_lower for x in ['meaning', 'nghia', 'dich']):
            meaning_col = idx + 1
    
    print(f"\n✅ Cột Test: {headers[test_col-1] if test_col else 'Không tìm thấy'} (Cột {test_col if test_col else 'N/A'})")
    print(f"✅ Cột Part: {headers[part_col-1] if part_col else 'Không tìm thấy'} (Cột {part_col if part_col else 'N/A'})")
    print(f"✅ Cột Word/Từ vựng: {headers[word_col-1] if word_col else 'Không tìm thấy'} (Cột {word_col if word_col else 'N/A'})")
    print(f"✅ Cột Meaning/Nghĩa: {headers[meaning_col-1] if meaning_col else 'Không tìm thấy'} (Cột {meaning_col if meaning_col else 'N/A'})")
    
    # Đếm số test và part
    if test_col and part_col:
        tests = set()
        parts = set()
        for row_idx in range(2, sheet.max_row + 1):
            test_val = sheet.cell(row=row_idx, column=test_col).value
            part_val = sheet.cell(row=row_idx, column=part_col).value
            if test_val:
                tests.add(str(test_val).strip())
            if part_val:
                parts.add(str(part_val).strip())
        
        print(f"\n📊 Thống kê:")
        print(f"   - Số Test khác nhau: {len(tests)}")
        if len(tests) <= 10:
            print(f"   - Danh sách Test: {', '.join(sorted(tests))}")
        print(f"   - Số Part khác nhau: {len(parts)}")
        if len(parts) <= 10:
            print(f"   - Danh sách Part: {', '.join(sorted(parts))}")
    
    print("\n" + "=" * 80)
    print("✅ Hoàn tất đọc file!")
    print("=" * 80)
    
except Exception as e:
    print(f"❌ Lỗi khi đọc file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)


